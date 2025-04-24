import requests
import time
import json

# Liste pour stocker les résultats
results = []

def log_message(message, decoration="-"):
    print(f"{decoration * 50}\n{message}\n{decoration * 50}")

def initialiser_conversation(company_id):
    log_message("Initialisation de la conversation", "=")
    
    base_url = "https://directline.asksuite.com/directline/conversations"
    params = {
        "isTest": "false",
        "companyId": company_id
    }
    
    print("🔄 Initialisation de la conversation...")
    response = requests.post(base_url, params=params)
    
    if response.status_code == 200:
        resp_body = response.json()
        conversation_id = resp_body.get("conversationId")
        if conversation_id:
            print(f"✅ Conversation initialisée avec succès (ID: {conversation_id})")
            return conversation_id
        else:
            print("❌ Impossible de récupérer l'ID de conversation")
            return None
    else:
        print(f"❌ Échec de l'initialisation (Statut : {response.status_code})")
        print(f"Réponse : {response.text}")
        return None

def interroger_chatbot(nom_hotel, company_id, payload, question_text):
    log_message(f"Interrogation du chatbot pour : {nom_hotel}", "=")
    
    # Initialiser la conversation
    conversation_id = initialiser_conversation(company_id)
    if not conversation_id:
        print("❌ Impossible de continuer sans ID de conversation")
        return
    
    # Construire l'URL avec le conversation_id
    url = f"https://directline.asksuite.com/directline/conversations/{conversation_id}/activities"
    
    headers = {"Content-Type": "application/json"}
    payload["text"] = question_text
    
    print(f"\n➡️ Envoi du message : \"{question_text}\" à {nom_hotel}...\n")
    response = requests.post(url, json=payload, headers=headers)
    
    if response.status_code == 200:
        print("✅ Message envoyé avec succès !")
    else:
        print(f"❌ Échec de l'envoi du message ! (Statut : {response.status_code})")
        print(f"Réponse : {response.text}")
        return
    
    print("⏳ En attente de la réponse du chatbot...\n")
    time.sleep(5)
    
    print("📨 Récupération de la réponse du chatbot...\n")
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        resp_body = response.json()
        
        if "activities" in resp_body and resp_body["activities"]:
            dernier_message = resp_body["activities"][-1].get("text", "Aucun texte disponible")
            
            if dernier_message == question_text:
                print("⚠️ Aucune réponse reçue du chatbot.")
            else:
                results.append((nom_hotel, question_text, dernier_message))
                print(f"💬 Réponse du chatbot : \"{dernier_message}\"\n")
        else:
            results.append((nom_hotel, question_text, "Échec de l'envoi"))
            print("⚠️ Aucune activité trouvée dans la réponse.")
    else:
        print(f"❌ Échec de la récupération de la réponse ! (Statut : {response.status_code})")
        results.append((nom_hotel, question_text, "Échec de la récupération"))
        print(f"Réponse : {response.text}\n")
    
    log_message(f"Fin de la session pour {nom_hotel}", "=")

donnees_hotels = [
    ("Lani Suites", "lanis-suites-de-luxe", {
        "event": None,
        "type": "message",
        "text": "",
        "from": {
            "id": "lanis-suites-de-luxe",
            "name": "Visitante",
            "language": "en-us"
        },
        "locale": "en-us",
        "textFormat": "plain",
        "timestamp": "2025-04-02T19:16:40.647Z",
        "channelData": {
            "clientActivityId": "1743621186456.3476015860815116.2"
        }
    }, "Les serviettes de piscine sont-elles fournies ?"),
    
    ("Apartahotel Jardines de Sabatini", "apartahotel-jardines-de-sabatini", {
        "event": None,
        "type": "message",
        "text": "",
        "from": {
            "id": "apartahotel-jardines-de-sabatini",
            "name": "Visitante",
            "language": "en-us"
        },
        "locale": "en-us",
        "textFormat": "plain",
        "timestamp": "2025-04-02T19:29:55.646Z",
        "channelData": {
            "clientActivityId": "1743621258740.49060530311373696.4"
        }
    }, "Quels soins de spa proposez-vous, s'il vous plaît ?"),
    
    ("Hard Rock Hotel", "hard-rock-cafe-new-york", {
        "event": None,
        "type": "message",
        "text": "",
        "from": {
            "id": "hard-rock-cafe-new-york",
            "name": "Visitante",
            "language": "en-us"
        },
        "locale": "en-us",
        "textFormat": "plain",
        "timestamp": "2025-04-02T19:32:05.759Z",
        "channelData": {
            "clientActivityId": "1743621374586.6982734224460425.0"
        },
        "entities": [
            {
                "type": "ClientCapabilities",
                "requiresBotState": True,
                "supportsTts": True,
                "supportsListening": True
            }
        ]
    }, "Je souhaite faire une réservation")
]

for hotel in donnees_hotels:
    interroger_chatbot(*hotel)
    
# Générer un fichier Excel
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# Création du DataFrame
df = pd.DataFrame(results, columns=["Hôtel", "Question", "Réponse"])

# Nom du fichier Excel
excel_filename = "asksuite_chatbot_responses.xlsx"

# Sauvegarde dans un fichier Excel
df.to_excel(excel_filename, index=False, engine='openpyxl')

# Chargement du fichier Excel pour le formatage
wb = Workbook()
ws = wb.active
ws.title = "Réponses Chatbot"

# Ajout des en-têtes avec style
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
border_style = Border(left=Side(style="thin"), 
                      right=Side(style="thin"), 
                      top=Side(style="thin"), 
                      bottom=Side(style="thin"))

for col_num, col_name in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border_style

# Remplissage des données et application du style
alternate_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for row_num, row in enumerate(df.values, 2):
    for col_num, value in enumerate(row, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.border = border_style
        if row_num % 2 == 0:
            cell.fill = alternate_fill  # Couleur alternée pour les lignes paires

# Ajustement automatique de la largeur des colonnes
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Obtenir la lettre de la colonne
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2  # Ajustement

# Sauvegarde du fichier formaté
wb.save(excel_filename)

print(f"✅ Fichier Excel généré avec succès : {excel_filename}")
